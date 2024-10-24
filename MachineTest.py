import openpyxl
from openpyxl import Workbook
import os

# File to store user data
file_name = "users.xlsx"

# Function to add a user
def add_user():
    name = input("Enter Name: ")
    email = input("Enter Email: ")
    phone = input("Enter Phone Number: ")

    # Check if the Excel file exists
    if not os.path.exists(file_name):
        # Create a new Excel file with headers
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email", "Phone Number"])  # Column headers
        wb.save(file_name)

    # Load the existing Excel file
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # Append the new user data
    ws.append([name, email, phone])
    wb.save(file_name)
    print("User added successfully!\n")

# Function to display users
def display_users():
    # Check if the file exists
    if not os.path.exists(file_name):
        print("No users found. Add users first.\n")
        return

    # Load the workbook and read data
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # Iterate through the rows and display the users
    print("Stored Users:")
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
        print(f"Name: {row[0]}, Email: {row[1]}, Phone: {row[2]}")
    print()

# Main function to handle user input
def main():
    while True:
        print("Make a choice to proceed:")
        print("Press '1' to Add user.")
        print("Press '2' to Display users.")
        print("Press '3' to Exit.")
        choice = input("Press any key to continue: ")

        if choice == '1':
            add_user()
        elif choice == '2':
            display_users()
        elif choice == '3':
            print("Exiting the program.")
            break
        else:
            print("Invalid choice. Please try again.\n")

if __name__ == "__main__":
    main()
